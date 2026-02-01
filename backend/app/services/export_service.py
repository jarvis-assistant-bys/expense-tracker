from pathlib import Path
from datetime import date, datetime
from typing import List
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from weasyprint import HTML, CSS

from app.models.expense import Expense

class ExportService:
    def __init__(self, output_dir: Path = Path("exports")):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_excel(self, expenses: List[Expense], month: int, year: int) -> io.BytesIO:
        """Génère un fichier Excel avec les dépenses du mois."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Note de frais {month:02d}/{year}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = ["Date", "Description", "Catégorie", "Fournisseur", "HT (€)", "TVA (€)", "TTC (€)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Données
        total_ht = 0
        total_tva = 0
        total_ttc = 0
        
        for row, expense in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=expense.date.strftime("%d/%m/%Y") if expense.date else "").border = border
            ws.cell(row=row, column=2, value=expense.description or "").border = border
            ws.cell(row=row, column=3, value=expense.category or "").border = border
            ws.cell(row=row, column=4, value=expense.vendor or "").border = border
            ws.cell(row=row, column=5, value=expense.amount_ht or 0).border = border
            ws.cell(row=row, column=6, value=expense.tva or 0).border = border
            ws.cell(row=row, column=7, value=expense.amount_ttc or 0).border = border
            
            total_ht += expense.amount_ht or 0
            total_tva += expense.tva or 0
            total_ttc += expense.amount_ttc or 0
        
        # Ligne totaux
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=round(total_ht, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=round(total_tva, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=round(total_ttc, 2)).font = Font(bold=True)
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Sauvegarder en mémoire
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_pdf(self, expenses: List[Expense], month: int, year: int, 
                     name: str = "Thomas Belardy") -> io.BytesIO:
        """Génère un PDF de note de frais."""
        
        # Calculer totaux
        total_ht = sum(e.amount_ht or 0 for e in expenses)
        total_tva = sum(e.tva or 0 for e in expenses)
        total_ttc = sum(e.amount_ttc or 0 for e in expenses)
        
        # Générer les lignes du tableau
        rows_html = ""
        for expense in expenses:
            rows_html += f"""
            <tr>
                <td>{expense.date.strftime("%d/%m/%Y") if expense.date else ""}</td>
                <td>{expense.description or ""}</td>
                <td>{expense.category or ""}</td>
                <td>{expense.vendor or ""}</td>
                <td class="number">{expense.amount_ht or 0:.2f} €</td>
                <td class="number">{expense.tva or 0:.2f} €</td>
                <td class="number">{expense.amount_ttc or 0:.2f} €</td>
            </tr>
            """
        
        month_names = {
            1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
            5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
            9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
        }
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 40px;
                    color: #333;
                }}
                h1 {{
                    color: #2c3e50;
                    border-bottom: 3px solid #3498db;
                    padding-bottom: 10px;
                }}
                .header {{
                    margin-bottom: 30px;
                }}
                .meta {{
                    color: #666;
                    margin-bottom: 20px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th {{
                    background-color: #3498db;
                    color: white;
                    padding: 12px 8px;
                    text-align: left;
                }}
                td {{
                    padding: 10px 8px;
                    border-bottom: 1px solid #ddd;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                .number {{
                    text-align: right;
                }}
                .total-row {{
                    font-weight: bold;
                    background-color: #ecf0f1 !important;
                    border-top: 2px solid #3498db;
                }}
                .footer {{
                    margin-top: 40px;
                    padding-top: 20px;
                    border-top: 1px solid #ddd;
                    color: #666;
                    font-size: 0.9em;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Note de Frais</h1>
                <div class="meta">
                    <p><strong>Nom :</strong> {name}</p>
                    <p><strong>Période :</strong> {month_names[month]} {year}</p>
                    <p><strong>Date d'édition :</strong> {datetime.now().strftime("%d/%m/%Y")}</p>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Description</th>
                        <th>Catégorie</th>
                        <th>Fournisseur</th>
                        <th class="number">HT</th>
                        <th class="number">TVA</th>
                        <th class="number">TTC</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                    <tr class="total-row">
                        <td colspan="4">TOTAL</td>
                        <td class="number">{total_ht:.2f} €</td>
                        <td class="number">{total_tva:.2f} €</td>
                        <td class="number">{total_ttc:.2f} €</td>
                    </tr>
                </tbody>
            </table>
            
            <div class="footer">
                <p>Document généré automatiquement par Expense Tracker</p>
            </div>
        </body>
        </html>
        """
        
        # Générer le PDF
        output = io.BytesIO()
        HTML(string=html_content).write_pdf(output)
        output.seek(0)
        return output

export_service = ExportService()
